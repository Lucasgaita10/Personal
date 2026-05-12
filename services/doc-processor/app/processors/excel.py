"""Excel/CSV processor — extracts each sheet as text + retains numeric tables."""
from __future__ import annotations

from dataclasses import dataclass
import io

import openpyxl
import pandas as pd


@dataclass
class SheetContent:
    sheet: str
    text: str
    rows: int
    cols: int
    table: list[list[str]]


def extract_xlsx(buf: bytes) -> list[SheetContent]:
    out: list[SheetContent] = []
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True, read_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        if not rows:
            continue
        # Plain-text projection useful for embeddings
        txt = "\n".join("\t".join(r) for r in rows)
        out.append(
            SheetContent(
                sheet=name,
                text=f"# Sheet: {name}\n{txt}",
                rows=len(rows),
                cols=max(len(r) for r in rows),
                table=rows[:200],  # cap to keep DB sane
            )
        )
    return out


def extract_csv(buf: bytes) -> SheetContent:
    df = pd.read_csv(io.BytesIO(buf))
    rows = [list(map(str, df.columns))] + df.astype(str).values.tolist()
    txt = "\n".join("\t".join(r) for r in rows)
    return SheetContent(sheet="csv", text=txt, rows=len(rows), cols=len(rows[0]) if rows else 0, table=rows[:200])
